import csv
import io

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..processes.shared import CSV, FISHERY


def adjust_column_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column  # Get the column name (A, B, C, etc.)
        column_letter = get_column_letter(column)  # Convert numeric index to letter
        
        # Find the maximum length of the content in each column
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Set the column width to slightly larger than the max length (for padding)
        adjusted_width = max_length + 2  # Add extra space for padding
        ws.column_dimensions[column_letter].width = adjusted_width


def get_sheets(feature):
    """ Returns an ordered dictionary of CSV results """
    results = {}
    for name, func in FISHERY.items():
        results[name] = func(feature, CSV)['Report']
    return results


def return_xlsx(feature):
    # create Excel content, add tables, stream result
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb["Sheet"])

    for name, content in get_sheets(feature).items():
        ws = wb.create_sheet(name)
        wb.active = ws
        # Parse the CSV data and write to the Excel sheet
        csv_reader = csv.reader(content.splitlines())
        for row_idx, row in enumerate(csv_reader, start=1):
            try:
                if 'p style' in row[0]:
                    # Ignore fisheries with no data (which return HTML in first column of first row with data)
                    break
            except:
                pass
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        adjust_column_width(ws)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)

    wb.active = wb[wb.sheetnames[0]]
    wb.save(output)

    # Return binary response
    output.seek(0)  # Reset stream pointer to the beginning
    return output.read()
